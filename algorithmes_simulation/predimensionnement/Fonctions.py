# -*- coding: utf-8 -*-
"""
Created on Thu Oct 18 15:18:20 2018

@author: sbouaraba
"""



import numpy as np
import matplotlib.pyplot as plt
import math
import copy
from scipy.optimize import fsolve

from predimensionnement import Chemin,Pertes, pertes_singulieres
from couts import Cost
from classes import Pompe, utilite, Echangeurs, piping

###########################Variable globale###############################

T_ambiant=-21

############################Fonction nécessaire au dev#########################

def selection(L_ssFlux,refssFlux):
    #fonction donnant le sous flux à partir de la référence dans le couple
    #ATTENTION S'APPLIQUE UNIQUEMENT SUR LES LISTES DE SOUS FLUX CHAUD OU FROID
    for ssflux in L_ssFlux:
        if ssflux.refFlux==refssFlux:
            return ssflux

def selection_couple(L_couple,ssflux):
    #fonction donnant le couple à partir du sous flux
    if ssflux.typeFlux=="c":
        for Couple in L_couple:
            if ssflux.refFlux==Couple.ssFluxC:
                return Couple
    else:
        for Couple in L_couple:
            if ssflux.refFlux==Couple.ssFluxF:
                return Couple
    print("Le sous flux n'est dans aucun couple")
    return None

def selection_utilite(L_utilite,refssFlux,type_flux):
    # la fonction donne l'utilité associé au couple
    
    # on choisit le type de l'utilité en fonction du ss flux
    if type_flux=="f":
        type_util="c"
    else:
        type_util="f"
    
    # on cherche l'utilité
    for util in L_utilite:
        if util.typeUtil==type_util:
            if util.ssFlux==refssFlux:
                return util
    return utilite.utilite((0,0),0,"r")

def nbssflux(L_ssflux,refflux):
    #fonction donnant le nombre de sous-flux
    compteur=0
    for i in range(len(L_ssflux)):
        if L_ssflux[i].refFlux[0]==refflux:
            compteur+=1
    return compteur

def nbflux(L_ssflux):
    #fonction donnant le nombre de flux
    Nflux=[]
    for ssflux in L_ssflux:
        Nflux.append(ssflux.refFlux[0])
    return max(Nflux)+1

def tmpchaud(L_ssflux):
    #renvoie la liste des sous flux chaud
    tmpchaud=[]
    for ssflux in L_ssflux:
        if ssflux.typeFlux=="c":
            tmpchaud.append(ssflux)
    return tmpchaud

def tmpfroid(L_ssflux):
    #renvoie la liste des sous flux froid
    tmpfroid=[]
    for ssflux in L_ssflux:
        if ssflux.typeFlux=="f":
            tmpfroid.append(ssflux)
    return tmpfroid

def positionzero(L_ssflux,numflux):
    #fonction donnant la position de l'élement du flux numflux avec un zéro dans sa ref
    compteur=0
    for ssflux in L_ssflux:
        if ssflux.refFlux[0]==numflux and ssflux.refFlux[1]==0:
            return compteur
        compteur+=1

def verif(L):
    #fonction permettant de vérifier si le tri est correctement fait lorsqu'on concatène les listes
    tmp=[]
    for ssflux in L:
        tmp.append(ssflux.refFlux)
    return tmp

def verifliste(L):
    #fonction permettant de vérifier si le trie est bien fait dans la liste de liste
    n=len(L)
    tmp=[[] for i in range(n)]
    for k in range(n):
        for ssflux in L[k]:
            tmp[k].append(ssflux.refFlux)
    return tmp

def tri(L_ssflux):
    #fonction triant la liste en ordre croissant du nombre de flux et sous flux
    
    nbFlux=nbflux(L_ssflux)
    #on trie en fonction seulement du nombre de flux
    tmp=[[] for k in range(nbFlux)]
    for k in range(nbFlux):
        for ssflux in L_ssflux:
            if ssflux.refFlux[0]==k:
                tmp[k].append(ssflux)
    #print(verifliste(tmp))
    #fonction triant les sous liste de sous flux
    def tri_bis(L_ssflux,L,refflux):
        nbssFlux=nbssflux(L_ssflux,refflux)
        tmp=[]
        start=0
        while len(tmp)!=nbssFlux:
            for ssflux in L:
                if ssflux.refFlux[1]==start:
                    tmp.append(ssflux)
            start+=1
        return tmp
    #on trie chaque liste de sous flux
    for i in range(len(tmp)):
        tmp[i]=tri_bis(L_ssflux,tmp[i],i)
    #print(verifliste(tmp))
    L_trie=[]
    def somme_liste(L):
        tmp=[]
        for liste in L:
            tmp+=liste
        return tmp
    #on crée une unique liste triée
    L_trie=somme_liste(tmp)
    #print(verif(L_trie))
    return L_trie

def basedivision(L_ssflux,refflux):
    #fonction donnant le le premier sous flux divisé qui sera considéré comme la base
    #ATTENTION LA FONCTION S'APPLIQUE SEULEMENT SUR LES LISTES DE SOUS FLUX CHAUD OU FROID
    if refflux[1]==0:
        return selection(L_ssflux,refflux)
    else:
        if selection(L_ssflux,(refflux[0],refflux[1]-1)).estDivise==False:
            return selection(L_ssflux,refflux)
        else:
            return basedivision(L_ssflux,(refflux[0],refflux[1]-1))

def nbr_flux(L_ssflux_type):
    # fonction donnant le nombre de flux d'un type donné
    L = []
    nbr = []
    for ssflux in L_ssflux_type:
        L.append(ssflux.refFlux[0])
        if not L[-1] in nbr:
            nbr.append(L[-1])
    return len(nbr)
    
#################################Partie Piping#################################

def chemin_dimensionnement_piping(ssFluxC,ssFluxF,r_echelle,m,n):
    #La fonction prend en entrée un objet couple, le rapport d'échelle du plan, la hauteur du plan en pixel et la largeur du plan en pixel
    
    #La fonction calcule la longueur de la tuyauterie et le nombre de coude
    #pour le flux chaud et le flux froid
    #La fonction actualise la valeur du diamètre de la tuyauterie, l'épaisseur 
    #de tuyauterie, l'épaisseur de calorifuge, le cout de la tuyauterie, le 
    #cout du calorifuge, la baisse de température due au déplacement dans la 
    #tuyauterie
    
    ######Calcul de la distance et du nombre de coude pour les deux flux#######
    
    ssFluxC.longTuy,ssFluxC.nbCoude,ssFluxF.longTuy,ssFluxF.nbCoude,x_ech,y_ech=Chemin.Chemin.chemin(ssFluxC.x,ssFluxC.y,ssFluxF.x,ssFluxF.y,r_echelle,m,n)
    
    #on attribut la position finale du sous flux à partir de celle de l'échangeur
    ssFluxC.xFinEch,ssFluxC.yFinEch=x_ech,y_ech
    ssFluxF.xFinEch,ssFluxF.yFinEch=x_ech,y_ech
    
    #on intialise les valeurs des surlongueurs à 0 dans le cas ou l'échange du ss flux n'est pas le dernier
    longTuyC,nbCoudeC,longTuyF,nbCoudeF=0,0,0,0
    
    if ssFluxC.refEch==False:
        longTuyC,nbCoudeC=Chemin.Chemin.chemin_ssflux(ssFluxC.xFinEch,ssFluxC.yFinEch,ssFluxC.objFlux.xFin,ssFluxC.objFlux.yFin,r_echelle,m,n)
    if ssFluxF.refEch==False:
        longTuyF,nbCoudeF=Chemin.Chemin.chemin_ssflux(ssFluxF.xFinEch,ssFluxF.yFinEch,ssFluxF.objFlux.xFin,ssFluxF.objFlux.yFin,r_echelle,m,n)
            
    #on ajoute la longueur due à la fin du tronçon
    ssFluxC.longTuy,ssFluxC.nbCoude,ssFluxF.longTuy,ssFluxF.nbCoude=ssFluxC.longTuy+longTuyC,ssFluxC.nbCoude+nbCoudeC,ssFluxF.longTuy+longTuyF,ssFluxF.nbCoude+nbCoudeF
    
    ####################Calcul pour le sous-flux chaud#########################
    
    #Calcul du premier diametre et epaisseur de tuyauterie
    Liste_diam_ep=piping.piping.section_droite(ssFluxC.debVol,ssFluxC.press,ssFluxC.rho)
    #Choix de la dimension normalisée
    [Dt,e,DN]=piping.piping.choix_taille(Liste_diam_ep)
    #Hypothèse de l'isolant utilisé avec une conductivité thermique de 0.04 W/m/K
    lamb_iso=0.04
    #Hypothèse la température ambiante est de 20°C
#    T_ambiant=-10
    #Calcul de l'épaisseur d'isolant en arrondissant au millimètre supérieur
    # on choisit la température max pour le calcul de l'isolant
    T_fluideC=max(ssFluxC.Te,ssFluxC.Ts)
    epIso=np.ceil(1000*piping.piping.epaisseur_min(lamb_iso,Dt/2+e,T_fluideC+273.15,T_ambiant+273.15))/1000
    #Actualise la valeur du piping dans l'objet ssFluxC
    ssFluxC.objPip=piping.piping(Dt,e,epIso,DN)    
    
    #Calcul du cout de la tuyauterie section droite
    cout_tuyauterie_droite=piping.piping.fitting(Dt)*ssFluxC.longTuy
    #Calcul cout de la tuyauterie coude avec hypothèse que le prix est 1.5 fois plus élevé que la section droite
    cout_tuyauterie_coude=piping.piping.fitting(Dt)*1.5*ssFluxC.nbCoude
    #Calcul du cout total de la tuyauterie
    cout_tuyauterie=cout_tuyauterie_droite+cout_tuyauterie_coude
    #Calcul du coût d'isolant
    cout_isolant=piping.piping.cout_calo_V(Dt/2+e,epIso)*ssFluxC.longTuy
    #Actualise la valeur des couts dans l'objet piping
    ssFluxC.objPip.coutTuy=cout_tuyauterie
    ssFluxC.objPip.coutIso=cout_isolant
    
    ####################Calcul pour le sous-flux froid#########################
    
    #Calcul du premier diametre et epaisseur de tuyauterie
    Liste_diam_ep=piping.piping.section_droite(ssFluxF.debVol,ssFluxF.press,ssFluxF.rho)
    #Choix de la dimension normalisée
    [Dt,e,DN]=piping.piping.choix_taille(Liste_diam_ep)
    #Hypothèse de l'isolant utilisé avec une conductivité thermique de 0.04 W/m/K
    lamb_iso=0.04
    #Hypothèse la température ambiante est de 20°C
#    T_ambiant=-10
    #Calcul de l'épaisseur d'isolant en arrondissant au millimètre supérieur
    # on choisit la température max pour le calcul de l'isolant
    T_fluideF=max(ssFluxF.Te,ssFluxF.Ts)
    epIso=np.ceil(1000*piping.piping.epaisseur_min(lamb_iso,Dt/2+e,T_fluideF+273.15,T_ambiant+273.15))/1000
    #Actualise la valeur du piping dans l'objet ssFluxF
    ssFluxF.objPip=piping.piping(Dt,e,epIso,DN)    
    
    #Calcul du cout de la tuyauterie section droite
    cout_tuyauterie_droite=piping.piping.fitting(Dt)*ssFluxF.longTuy
    #Calcul cout de la tuyauterie coude avec hypothèse que le prix est 1.5 fois plus élevé que la section droite
    cout_tuyauterie_coude=piping.piping.fitting(Dt)*1.5*ssFluxF.nbCoude
    #Calcul du cout total de la tuyauterie
    cout_tuyauterie=cout_tuyauterie_droite+cout_tuyauterie_coude
    #Calcul du coût d'isolant
    cout_isolant=piping.piping.cout_calo_V(Dt/2+e,epIso)*ssFluxF.longTuy
    #Actualise la valeur des couts dans l'objet piping
    ssFluxF.objPip.coutTuy=cout_tuyauterie
    ssFluxF.objPip.coutIso=cout_isolant

    #La fonction ne renvoie aucune valeur
    return None

def chemin_dimensionnement_piping_utilite(ssFluxC,r_echelle,m,n):
    #La fonction prend en entrée un sous flux, le rapport d'échelle du plan, la hauteur du plan en pixel et la largeur du plan en pixel
    
    #La fonction calcule la longueur de la tuyauterie et le nombre de coude
    #pour le flux avec une utilite jusqu'à la fin du tronçon
    #La fonction actualise la valeur du diamètre de la tuyauterie, l'épaisseur 
    #de tuyauterie, l'épaisseur de calorifuge, le cout de la tuyauterie, le 
    #cout du calorifuge, la baisse de température due au déplacement dans la 
    #tuyauterie
    
    #on intialise les valeurs des surlongueurs à 0 dans le cas ou l'échange du ss flux n'est pas le dernier
    longTuyC,nbCoudeC=0,0
    
    if ssFluxC.refEch==False:
        longTuyC,nbCoudeC=Chemin.Chemin.chemin_ssflux(ssFluxC.x,ssFluxC.y,ssFluxC.xFinEch,ssFluxC.yFinEch,r_echelle,m,n)
    else:
        print("ABUSE")
    #on ajoute la longueur due à la fin du tronçon
    ssFluxC.longTuy,ssFluxC.nbCoude=longTuyC,nbCoudeC
    
    ####################Calcul pour le sous-flux chaud#########################
    
    #Calcul du premier diametre et epaisseur de tuyauterie
    Liste_diam_ep=piping.piping.section_droite(ssFluxC.debVol,ssFluxC.press,ssFluxC.rho)
    #Choix de la dimension normalisée
    [Dt,e,DN]=piping.piping.choix_taille(Liste_diam_ep)
    #Hypothèse de l'isolant utilisé avec une conductivité thermique de 0.04 W/m/K
    lamb_iso=0.04
    #Hypothèse la température ambiante est de 20°C
#    T_ambiant=-10
    #Calcul de l'épaisseur d'isolant en arrondissant au millimètre supérieur
    # on choisit la température max pour le calcul de l'isolant
    T_fluide=max(ssFluxC.Te,ssFluxC.Ts)
    epIso=np.ceil(1000*piping.piping.epaisseur_min(lamb_iso,Dt/2+e,T_fluide+273.15,T_ambiant+273.15))/1000
    #Actualise la valeur du piping dans l'objet ssFluxC
    ssFluxC.objPip=piping.piping(Dt,e,epIso,DN)    
    
    #Calcul du cout de la tuyauterie section droite
    cout_tuyauterie_droite=piping.piping.fitting(Dt)*ssFluxC.longTuy
    #Calcul cout de la tuyauterie coude avec hypothèse que le prix est 1.5 fois plus élevé que la section droite
    cout_tuyauterie_coude=piping.piping.fitting(Dt)*1.5*ssFluxC.nbCoude
    #Calcul du cout total de la tuyauterie
    cout_tuyauterie=cout_tuyauterie_droite+cout_tuyauterie_coude
    #Calcul du coût d'isolant
    cout_isolant=piping.piping.cout_calo_V(Dt/2+e,epIso)*ssFluxC.longTuy
    #Actualise la valeur des couts dans l'objet piping
    ssFluxC.objPip.coutTuy=cout_tuyauterie
    ssFluxC.objPip.coutIso=cout_isolant

    #La fonction ne renvoie aucune valeur
    return None

def chemin_dimensionnement_piping_position_done(ssFluxC,ssFluxF,r_echelle,m,n):
    #La fonction prend en entrée un objet couple, le rapport d'échelle du plan, la hauteur du plan en pixel et la largeur du plan en pixel
    
    #La fonction calcule la longueur de la tuyauterie et le nombre de coude
    #pour le flux chaud et le flux froid
    #La fonction actualise la valeur du diamètre de la tuyauterie, l'épaisseur 
    #de tuyauterie, l'épaisseur de calorifuge, le cout de la tuyauterie, le 
    #cout du calorifuge, la baisse de température due au déplacement dans la 
    #tuyauterie
    
    ######Calcul de la distance et du nombre de coude pour les deux flux#######
    
    ssFluxC.longTuy,ssFluxC.nbCoude,ssFluxF.longTuy,ssFluxF.nbCoude,x_ech,y_ech=Chemin.Chemin.chemin(ssFluxC.x,ssFluxC.y,ssFluxF.x,ssFluxF.y,r_echelle,m,n)
    
    #on attribut la position finale du sous flux à partir de celle de l'échangeur
#    ssFluxC.xFinEch,ssFluxC.yFinEch=x_ech,y_ech
#    ssFluxF.xFinEch,ssFluxF.yFinEch=x_ech,y_ech
    
    #on intialise les valeurs des surlongueurs à 0 dans le cas ou l'échange du ss flux n'est pas le dernier
    longTuyC,nbCoudeC,longTuyF,nbCoudeF=0,0,0,0
    
    if ssFluxC.refEch==False:
        longTuyC,nbCoudeC=Chemin.Chemin.chemin_ssflux(ssFluxC.xFinEch,ssFluxC.yFinEch,ssFluxC.objFlux.xFin,ssFluxC.objFlux.yFin,r_echelle,m,n)
    if ssFluxF.refEch==False:
        longTuyF,nbCoudeF=Chemin.Chemin.chemin_ssflux(ssFluxF.xFinEch,ssFluxF.yFinEch,ssFluxF.objFlux.xFin,ssFluxF.objFlux.yFin,r_echelle,m,n)
    
    #on gère le fait que le flux peut-être divisé il sera alors rattaché au premier à l'être
    longTuyCDiv,nbCoudeCDiv,longTuyFDiv,nbCoudeFDiv=0,0,0,0
    
    if ssFluxF.estDivise==True:
        longTuyFDiv,nbCoudeFDiv=Chemin.Chemin.chemin_ssflux(ssFluxF.xFinEch,ssFluxF.yFinEch,ssFluxF.xFin,ssFluxF.yFin,r_echelle,m,n)
    if ssFluxC.estDivise==True:
        longTuyCDiv,nbCoudeCDiv=Chemin.Chemin.chemin_ssflux(ssFluxC.xFinEch,ssFluxC.yFinEch,ssFluxC.xFin,ssFluxC.yFin,r_echelle,m,n)
    
    #on ajoute la longueur due à la fin du tronçon & à la division
    ssFluxC.longTuy,ssFluxC.nbCoude,ssFluxF.longTuy,ssFluxF.nbCoude=ssFluxC.longTuy+longTuyC+longTuyCDiv,ssFluxC.nbCoude+nbCoudeC+nbCoudeCDiv,ssFluxF.longTuy+longTuyF+longTuyFDiv,ssFluxF.nbCoude+nbCoudeF+nbCoudeFDiv
    
    # on actualise la valeur des longueurs de tuyauterie division/fin
    ssFluxC.longTuyFin,ssFluxC.longTuyDiv,ssFluxF.longTuyFin,ssFluxF.longTuyDiv=longTuyC,longTuyCDiv,longTuyF,longTuyFDiv
    
    ####################Calcul pour le sous-flux chaud#########################
    
    #Calcul du premier diametre et epaisseur de tuyauterie
    Liste_diam_ep=piping.piping.section_droite(ssFluxC.debVol,ssFluxC.press,ssFluxC.rho)
    #Choix de la dimension normalisée
    [Dt,e,DN]=piping.piping.choix_taille(Liste_diam_ep)
    #Hypothèse de l'isolant utilisé avec une conductivité thermique de 0.04 W/m/K
    lamb_iso=0.04
    #Hypothèse la température ambiante est de 20°C
#    T_ambiant=20
    #Calcul de l'épaisseur d'isolant en arrondissant au millimètre supérieur
    # on choisit la température max pour le calcul de l'isolant
    T_fluideC=max(ssFluxC.Te,ssFluxC.Ts)
    epIso=np.ceil(1000*piping.piping.epaisseur_min(lamb_iso,Dt/2+e,T_fluideC+273.15,T_ambiant+273.15))/1000
    #Actualise la valeur du piping dans l'objet ssFluxC
    ssFluxC.objPip=piping.piping(Dt,e,epIso,DN)    
    
    #Calcul du cout de la tuyauterie section droite
    cout_tuyauterie_droite=piping.piping.fitting(Dt)*ssFluxC.longTuy
    #Calcul cout de la tuyauterie coude avec hypothèse que le prix est 1.5 fois plus élevé que la section droite
    cout_tuyauterie_coude=piping.piping.fitting(Dt)*1.5*ssFluxC.nbCoude
    #Calcul du cout total de la tuyauterie
    cout_tuyauterie=cout_tuyauterie_droite+cout_tuyauterie_coude
    #Calcul du coût d'isolant
    cout_isolant=piping.piping.cout_calo_V(Dt/2+e,epIso)*ssFluxC.longTuy
    #Actualise la valeur des couts dans l'objet piping
    ssFluxC.objPip.coutTuy=cout_tuyauterie
    ssFluxC.objPip.coutIso=cout_isolant
    
    ####################Calcul pour le sous-flux froid#########################
    
    #Calcul du premier diametre et epaisseur de tuyauterie
    Liste_diam_ep=piping.piping.section_droite(ssFluxF.debVol,ssFluxF.press,ssFluxF.rho)
    #Choix de la dimension normalisée
    [Dt,e,DN]=piping.piping.choix_taille(Liste_diam_ep)
    #Hypothèse de l'isolant utilisé avec une conductivité thermique de 0.04 W/m/K
    lamb_iso=0.04
    #Hypothèse la température ambiante est de 20°C
#    T_ambiant=20
    #Calcul de l'épaisseur d'isolant en arrondissant au millimètre supérieur
    # on choisit la température max pour le calcul de l'isolant
    T_fluideF=max(ssFluxF.Te,ssFluxF.Ts)
    epIso=np.ceil(1000*piping.piping.epaisseur_min(lamb_iso,Dt/2+e,T_fluideF+273.15,T_ambiant+273.15))/1000
    #Actualise la valeur du piping dans l'objet ssFluxF
    ssFluxF.objPip=piping.piping(Dt,e,epIso,DN)    
    
    #Calcul du cout de la tuyauterie section droite
    cout_tuyauterie_droite=piping.piping.fitting(Dt)*ssFluxF.longTuy
    #Calcul cout de la tuyauterie coude avec hypothèse que le prix est 1.5 fois plus élevé que la section droite
    cout_tuyauterie_coude=piping.piping.fitting(Dt)*1.5*ssFluxF.nbCoude
    #Calcul du cout total de la tuyauterie
    cout_tuyauterie=cout_tuyauterie_droite+cout_tuyauterie_coude
    #Calcul du coût d'isolant
    cout_isolant=piping.piping.cout_calo_V(Dt/2+e,epIso)*ssFluxF.longTuy
    #Actualise la valeur des couts dans l'objet piping
    ssFluxF.objPip.coutTuy=cout_tuyauterie
    ssFluxF.objPip.coutIso=cout_isolant

    #La fonction ne renvoie aucune valeur
    return None

def perte_temp(ssFlux):
    #La fonction prend en entrée un objet couple, le rapport d'échelle du plan, la hauteur du plan en pixel et la largeur du plan en pixel

    #Calcul de la température à la fin de la tuyauterie
    Cp=ssFlux.CP/(ssFlux.debVol*ssFlux.rho)
    #Hypothèse sur la conductivité de la tuyauterie suppose acier carbone 50 W/m/K
#    T_ambiant=20
    Dt,e,epIso=ssFlux.objPip.diamTuy,ssFlux.objPip.epTuy,ssFlux.objPip.epIso
    lamb_iso=0.04
    
    ssFlux.Te=Pertes.Pertes_de_chaleur_linéaire(ssFlux.debVol,ssFlux.Te+273.15,T_ambiant+273.15,Dt,Dt+2*e,Dt+2*(e+epIso),ssFlux.longTuy-ssFlux.longTuyFin-ssFlux.longTuyDiv,Cp,ssFlux.rho,ssFlux.lamb,lamb_iso)
    # ssFlux.Te=Pertes.Pertes.Pertes_de_chaleur_linéaire(ssFlux.debVol,ssFlux.Te+273.15,T_ambiant+273.15,Dt,Dt+2*e,Dt+2*(e+epIso),ssFlux.longTuy-ssFlux.longTuyFin-ssFlux.longTuyDiv,Cp,ssFlux.rho,ssFlux.lamb,lamb_iso)

    #La fonction ne renvoie aucune valeur
    return None

def perte_temp_fin(ssFlux):
    #La fonction prend en entrée un objet couple, le rapport d'échelle du plan, la hauteur du plan en pixel et la largeur du plan en pixel

    #Calcul de la température à la fin de la tuyauterie
    Cp=ssFlux.CP/(ssFlux.debVol*ssFlux.rho)
    #Hypothèse sur la conductivité de la tuyauterie suppose acier carbone 50 W/m/K
#    T_ambiant=20
    Dt,e,epIso=ssFlux.objPip.diamTuy,ssFlux.objPip.epTuy,ssFlux.objPip.epIso
    lamb_iso=0.04
    
    ssFlux.Ts=Pertes.Pertes_de_chaleur_linéaire(ssFlux.debVol,ssFlux.Ts+273.15,T_ambiant+273.15,Dt,Dt+2*e,Dt+2*(e+epIso),ssFlux.longTuyFin,Cp,ssFlux.rho,ssFlux.lamb,lamb_iso)
    # ssFlux.Ts=Pertes.Pertes.Pertes_de_chaleur_linéaire(ssFlux.debVol,ssFlux.Ts+273.15,T_ambiant+273.15,Dt,Dt+2*e,Dt+2*(e+epIso),ssFlux.longTuyFin,Cp,ssFlux.rho,ssFlux.lamb,lamb_iso)

    #La fonction ne renvoie aucune valeur
    return None

def perte_reso(ssflux):
    # la fonction donne la puissance nécessaire de l'échangeur de chaleur
    
    # on donne la température recherchée
    Ts=ssflux.Ts
    
    #Calcul de la température à la fin de la tuyauterie
    Cp=ssflux.CP/(ssflux.debVol*ssflux.rho)
    #Hypothèse sur la conductivité de la tuyauterie suppose acier carbone 50 W/m/K
#    T_ambiant=20
    Dt,e,epIso=ssflux.objPip.diamTuy,ssflux.objPip.epTuy,ssflux.objPip.epIso
    lamb_iso=0.04
    print(type(Dt),type(e),type(epIso))
    # on définit la fonction donnant l'équation à résoudre
    def f(Q):
        return Pertes.Pertes_de_chaleur_linéaire(ssflux.debVol,ssflux.Te+273.15-Q/ssflux.CP,T_ambiant+273.15,Dt,Dt+2*e,Dt+2*(e+epIso),ssflux.longTuyFin,Cp,ssflux.rho,ssflux.lamb,lamb_iso)-Ts
        # return Pertes.Pertes.Pertes_de_chaleur_linéaire(ssflux.debVol,ssflux.Te+273.15-Q/ssflux.CP,T_ambiant+273.15,Dt,Dt+2*e,Dt+2*(e+epIso),ssflux.longTuyFin,Cp,ssflux.rho,ssflux.lamb,lamb_iso)-Ts
    
    print(f(10000),type(f(10000)))
    # on résout l'équation ce qui nous donne la puissance de l'échangeur
    puissance_echangeur=fsolve(f,10000) # en W
    
    print(type(puissance_echangeur),puissance_echangeur)
    # on calcule aussi la température en sortie d'échangeur
    Ts=(ssflux.Te+273.15-puissance_echangeur/ssflux.CP)-273.15
    
    return puissance_echangeur/1000,Ts # en kW Ts en °C

def pertes_de_charge_piping(ssflux):
    #La fonction calcule les pertes de charge linéaire et singulière pour 
    #chaque flux et actualise leurs valeurs
    
    ##########################Calcul pour le flux chaud########################
    
    #Hypothèse sur la rugosité en m de l'acier
    k=0.05e-3
    #Calcul des pertes de charges régulière
    pertes_reg=Pertes.perte_lin_qv(ssflux.longTuy,ssflux.objPip.diamTuy,ssflux.debVol,k,ssflux.rho,ssflux.mu)
    # pertes_reg=Pertes.Pertes.perte_lin_qv(ssflux.longTuy,ssflux.objPip.diamTuy,ssflux.debVol,k,ssflux.rho,ssflux.mu)
#    print("la perte de charge regulière est de :",pertes_reg/1e5,"bar")
    #Calcul des pertes de charge singulière
    pertes_sing=ssflux.nbCoude*pertes_singulieres.perte_sin_qv("Coude_brusque",ssflux.objPip.diamTuy,ssflux.debVol,ssflux.rho,ssflux.mu,np.pi/2)
    # pertes_sing=ssflux.nbCoude*pertes_singulieres.pertes_singulieres.perte_sin_qv("Coude_brusque",ssflux.objPip.diamTuy,ssflux.debVol,ssflux.rho,ssflux.mu,np.pi/2)
#    print("la perte de charge singulière est de :",pertes_sing/1e5,"bar")
    #Actualise la valeur des pertes de charge dans l'objet ssFluxC
    ssflux.objPip.pertes=pertes_reg+pertes_sing
    
    return None

def chemin_perte_inital(ssflux,r_echelle,m,n):
    #la fonction donne la longueur du tracé initial et les pertes de charges qui en découle
    longTuy,nbCoude=Chemin.Chemin.chemin_ssflux(ssflux.objFlux.xDeb,ssflux.objFlux.yDeb,ssflux.objFlux.xFin,ssflux.objFlux.yFin,r_echelle,m,n)
    
    #Hypothèse sur la rugosité en m de l'acier
    k=0.05e-3
    #Calcul des pertes de charges régulière
    pertes_reg=Pertes.perte_lin_qv(longTuy,ssflux.objPip.diamTuy,ssflux.debVol,k,ssflux.rho,ssflux.mu)
    #  pertes_reg=Pertes.Pertes.perte_lin_qv(longTuy,ssflux.objPip.diamTuy,ssflux.debVol,k,ssflux.rho,ssflux.mu)
#    print("la perte de charge regulière est de :",pertes_reg/1e5,"bar")
    #Calcul des pertes de charge singulière
    pertes_sing=nbCoude*pertes_singulieres.perte_sin_qv("Coude_brusque",ssflux.objPip.diamTuy,ssflux.debVol,ssflux.rho,ssflux.mu,np.pi/2)
    # pertes_sing=nbCoude*pertes_singulieres.pertes_singulieres.perte_sin_qv("Coude_brusque",ssflux.objPip.diamTuy,ssflux.debVol,ssflux.rho,ssflux.mu,np.pi/2)
#    print("la perte de charge singulière est de :",pertes_sing/1e5,"bar")
    #Actualise la valeur des pertes de charge dans l'objet ssFlux
    dP=pertes_reg+pertes_sing
    
    return dP

###############################Partie echangeur################################

def classement_ech(ssfluxC,ssfluxF):
    #Classe plusieurs échangeurs en fonction de leur plage de fonctionnement (température,pression)
    
    #Creation BDD echangeur avec type/Tmin/Tmax/Pmin/Pmax
    #La BDD devra être complétée
    BDD=[["Tubular",-30,1100,0,1000e5],["Plated",-30,150,0,15e5]]
    #Initialise le classement
    classement=[]
    #Donne le nombre de type d'échangeur à tester
    longBDD=len(BDD)
    for k in range(longBDD):
        # limite de température
        T_min = min(ssfluxC.Te,ssfluxC.Ts,ssfluxF.Ts,ssfluxF.Te)
        T_max = max(ssfluxC.Te,ssfluxC.Ts,ssfluxF.Ts,ssfluxF.Te)
        # limite de pression
        P_min = min(ssfluxC.press,ssfluxF.press)
        P_max = max(ssfluxC.press,ssfluxF.press)
        #On vérifie si les conditions de pression et de température sont respecter pour les deux fluides du couple
        if T_min>=BDD[k][1] and T_max<=BDD[k][2] and P_min>=BDD[k][3] and P_max<=BDD[k][4] :
            classement.append(BDD[k][0])
    # renvoie une liste de string donnant le nom de l'échangeur sélectionné
    # print("len(classement)",len(classement))
    if len(classement)==0:
        classement=["Tubular","Plated"]
    return classement

def dimensionnement_echangeur(ssFluxC,ssFluxF,type_ech):
    #La fonction donne la surface d'échange et le prix pour chaque type d'échangeur
    
    Q_lim = 4
    
    #Dans le cas d'un échangeur tubulaire 1 passe calandre 2n passes tube
    if type_ech=="Tubular":
        #on calcule la capacité thermique massique du fluide chaud
        Cp_c=ssFluxC.CP/(ssFluxC.debVol*ssFluxC.rho)
        #on calcule la capacité thermique massique du fluide froid
        Cp_f=ssFluxF.CP/(ssFluxF.debVol*ssFluxF.rho)
        
        #on calcule la surface d'échange de l'échangeur ainsi que la longueur des tubes
        
        #Hypothèses
        D_calandre=1 #diamètre de la calandre fixe
        D_int=0.016 #diamètre intérieur des tubes fixe
        D_ext=0.02 #diamètre extérieur des tubes fixe
        L_tube=4.88 #longueur des tubes fixe
        lamb_tube=50 #conductivité des tubes fixe
        
        Q_max = max(ssFluxC.debVol,ssFluxF.debVol)
        Q_min = min(ssFluxC.debVol,ssFluxF.debVol)
        if Q_max%Q_lim!=0:
            n_div = int(Q_max//Q_lim)+1
        else:
            n_div = int(Q_max//Q_lim)
            
        if Q_max == ssFluxC.debVol :
            Qc_calc = Q_max/n_div
            Qf_calc = Q_min/n_div
        else:
            Qc_calc = Q_min/n_div
            Qf_calc = Q_max/n_div
        # print('TUBULAR: '+str(Q_max)+", "+str(Q_min)+", "+str(n_div)+", "+str(Q_lim)+", "+str(Cp_c)+", "+str(Cp_f)+", ")
            
        if Q_max > Q_lim :
            cout_tot = 0
            pompesplitchaud= []
            pompesplitfroid= []
            print("N_DIVVVVVVVVVVVVVVVVVV", n_div, ssFluxC.debVol, ssFluxF.debVol)
            for k in range(n_div):
                surfEch,longTub,dP_f,dP_c=Echangeurs.Tubular(Qc_calc,Qf_calc,Cp_c,Cp_f,ssFluxC.Te,ssFluxC.Ts,ssFluxF.Te,ssFluxF.Ts,D_int,D_ext,D_calandre,ssFluxC.mu,ssFluxF.mu,ssFluxC.rho,ssFluxF.rho,L_tube,lamb_tube,ssFluxC.lamb,ssFluxF.lamb)
                pompesplitchaud.append(Pompe.puissance_fonctionnement_normal(Qc_calc,dP_c))
                pompesplitfroid.append(Pompe.puissance_fonctionnement_normal(Qf_calc,dP_f))
                #Calcul du cout de l'échangeur
                cout_tot += Cost.Cout(surfEch,type_ech,ssFluxC,ssFluxF)
                
            objEch=Echangeurs.Echangeurs(type_ech,surfEch,cout_tot,dP_c,dP_f)
            objEch.div = True
            objEch.pompeSplitChaud = max(pompesplitchaud)
            objEch.pompeSplitFroid = max(pompesplitfroid)
        else:
            #on calcule la surface d'échange de l'échangeur ainsi que la longueur des tubes
            
            #le fluide chaud est dans la calandre
            surfEch,longTub,dP_f,dP_c=Echangeurs.Tubular(ssFluxC.debVol,ssFluxF.debVol,Cp_c,Cp_f,ssFluxC.Te,ssFluxC.Ts,ssFluxF.Te,ssFluxF.Ts,D_int,D_ext,D_calandre,ssFluxC.mu,ssFluxF.mu,ssFluxC.rho,ssFluxF.rho,L_tube,lamb_tube,ssFluxC.lamb,ssFluxF.lamb)
    #        print(dP_f,dP_c)
            #Calcul du cout de l'échangeur
            cout=Cost.Cout(surfEch,type_ech,ssFluxC,ssFluxF)
            #Création de l'objet échangeur dans le couple
            objEch=Echangeurs.Echangeurs(type_ech,surfEch,cout,dP_c,dP_f)
        
    elif type_ech=="Plated":
        #on calcule la capacité thermique massique du fluide chaud
        Cp_c=ssFluxC.CP/(ssFluxC.debVol*ssFluxC.rho)
        #on calcule la capacité thermique massique du fluide froid
        Cp_f=ssFluxF.CP/(ssFluxF.debVol*ssFluxF.rho)
        
        Q_max = max(ssFluxC.debVol,ssFluxF.debVol)
        Q_min = min(ssFluxC.debVol,ssFluxF.debVol)
        if Q_max%Q_lim!=0:
            n_div = int(Q_max//Q_lim)+1
        else:
            n_div = int(Q_max//Q_lim)
            
        if Q_max == ssFluxC.debVol :
            Qc_calc = Q_max/n_div
            Qf_calc = Q_min/n_div
        else:
            Qc_calc = Q_min/n_div
            Qf_calc = Q_max/n_div
        
        #on calcule la surface d'échange de l'échangeur
        if Q_max > Q_lim :
            cout_tot = 0
            pompesplitchaud= []
            pompesplitfroid= []
            for k in range(n_div):
                DPMAX_opti = 80
                
                surfEch,dP_c,dP_f,Phi=Echangeurs.Plated_1986(Qc_calc*3600,Qf_calc*3600,Cp_c,Cp_f,ssFluxC.lamb,ssFluxF.lamb,ssFluxC.Te,ssFluxC.Ts,ssFluxF.Te,ssFluxF.Ts,ssFluxF.mu*1000,ssFluxC.mu*1000,ssFluxC.rho,ssFluxF.rho,3e-5,DPMAX_opti)
                
                pompesplitchaud.append(Pompe.puissance_fonctionnement_normal(Qc_calc,dP_c))
                pompesplitfroid.append(Pompe.puissance_fonctionnement_normal(Qf_calc,dP_f))
                
                #Calcul du cout de l'échangeur
                cout_tot += Cost.Cout(surfEch,type_ech,ssFluxC,ssFluxF)
                
            objEch=Echangeurs.Echangeurs(type_ech,surfEch,cout_tot,dP_c,dP_f)
            objEch.div = True
            objEch.pompeSplitChaud = max(pompesplitchaud)
            objEch.pompeSplitFroid = max(pompesplitfroid)
        #Hypothèses
#        e=2e-3 #épaisseur de chaque plaque
#        lamb_inox=15 #conductivité thermique de l'inox
#        plis=0.2e-3
#        e_k=e/lamb_inox #rapport de l'épaisseur de la paroi et de la conductvité thermique de la paroi ici Inox
#        #einter=2e-3 #largeur d'un plis
#        L=1 #longueur de la plaque
#        l=500 #largeur de la plaque
#        alpha=30 #angle des plis
        
#        surfEch,dP_c,dP_f=Echangeurs.Plated(ssFluxC.debVol,ssFluxF.debVol,Cp_c,Cp_f,ssFluxC.lamb,ssFluxF.lamb,ssFluxC.Te,ssFluxC.Ts,ssFluxF.Te,ssFluxF.Ts,ssFluxF.mu,ssFluxC.mu,ssFluxC.rho,ssFluxF.rho,e_k,e,plis,l,L,alpha,ssFluxC.objPip.diamTuy,ssFluxF.objPip.diamTuy)
        else:
            DPMAX_opti = 80
            
            surfEch,dP_c,dP_f,Phi=Echangeurs.Plated_1986(ssFluxC.debVol*3600,ssFluxF.debVol*3600,Cp_c,Cp_f,ssFluxC.lamb,ssFluxF.lamb,ssFluxC.Te,ssFluxC.Ts,ssFluxF.Te,ssFluxF.Ts,ssFluxF.mu*1000,ssFluxC.mu*1000,ssFluxC.rho,ssFluxF.rho,3e-5,DPMAX_opti)
            #surfEch,dP_c,dP_f=Echangeurs.Plated(objCouple.ssFluxC.debVol,objCouple.ssFluxF.debVol,objCouple.puissE,e,einter,objCouple.ssFluxC.Te,objCouple.ssFluxC.Ts,objCouple.ssFluxF.Te,objCouple.ssFluxF.Ts,L,l,objCouple.ssFluxC.rho,objCouple.ssFluxF.rho,objCouple.ssFluxC.mu,objCouple.ssFluxF.mu,Cp_c,Cp_f,objCouple.ssFluxC.lamb,objCouple.ssFluxF.lamb,objCouple.ssFluxC.objPip.diamTuy,objCouple.ssFluxF.objPip.diamTuy)
    #        print(dP_c,dP_f)
            #Calcul du cout de l'échangeur
            cout=Cost.Cout(surfEch,type_ech,ssFluxC,ssFluxF)
            #création de l'objet échangeur dans le couple
            objEch=Echangeurs.Echangeurs(type_ech,surfEch,cout,dP_c,dP_f)
        
    return objEch

###############################Calcul des KPI################################

def opex_utilite_froide(puissance):
    # fonction donnant le cout en élec de ces utilités 
    
    # on définit le coefficient de performance de la pompe
    COP=3.5
    
    # la puissance élec à fournir
    W_elec=puissance/COP
    
    # on définit le cout de l'électricité pour 1 an
    facteur_utilisation=0.8 # facteur donnant le pourcentage de fonctionnement dans l'année
    C_kWh=0.155
    C_kW=C_kWh*8760*facteur_utilisation
    
    # cout de l'électricité pour la PAC
    return W_elec*C_kW

def opex_utilite_chaude(puissance):
    # fonction donnant le cout en gaz naturel pour le chauffage en considérant chaudière eau reseau + échangeur
    
    # on définit les rendements des deux composants
    eta_echangeur=1
    eta_chaudiere=0.85
    
    # on calcule l'énergie en gaz nat nécessaire
    Q_GN=puissance/(eta_echangeur*eta_chaudiere)
    
    # on définit le cout du gaz nat
    facteur_utilisation=0.8 # facteur donnant le pourcentage de fonctionnement dans l'année
    C_kWh=0.00595
    C_kW=C_kWh*8760*facteur_utilisation
    
    # cout de GN pour le chauffage
    return C_kW*Q_GN
    




class Fonctions:
    
    def Reynolds(U,D,rho,mu):
        # Q en m^3/s
        # D en m
        # rho en kg/m^3
        # mu en kg/(m.s)
        
        Re=(rho*U*D)/mu
                 
        return(Re)
                 
    def Prandtl(mu,Cp,k):
        # mu en kg/(m.s)
        # Cp en J/(kg.K)
        # k en W/(m.K)
        
        
        Pr=mu*Cp/k
                 
        return(Pr)
    
    def Nusselt_int(Re,Pr,D,L):
        Nu=4.36
        # D en m
        # L en m
        # A l'intérieur du/des tubes :
    #Correlation COLBURN avec Pr(0.7-100) & l/D>60
        if (3000<Re and  (L/D)>60)  : return 0.023*((Pr)**(1/3))*((Re)**(0.8))
        if (3000<Re and  (L/D)>10)  : return 0.023*((Pr)**(0.4))*((Re)**(0.8))
        #Correlation COLBURN avec l/D<60
        if (3000<Re and  (L/D)<60) : return 0.023*((Pr)**(1/3))*((D/2)**(0.8))*(1+((D)/L)**0.7)
#        #Régime laminaire avec Température de surface supposé constante   
#        if (Re<3000) : return Nu=3.66
                       
        return(Nu)
    
    def Nusselt_ext(Re,Pr):
    
    #A l'extérieur des tubes.
        if (0.4<Re and Re<4 )  : Nu=0.083*((Pr)**(1/3))*((Re)**(0.8))
        if (4<Re and Re<40 )  : Nu=0.911*((Pr)**(1/3))*((Re)**(0.385))
        if (40<Re and Re<4000 )  : Nu=0.683*((Pr)**(1/3))*((Re)**(0.466))
        if (4000<Re and Re<40000 )  : Nu=0.193*((Pr)**(1/3))*((Re)**(0.617))
        if (40000<Re)  : Nu=0.027*((Pr)**(1/3))*((Re)**(0.805))
#        Formule utilisée par Geoffrey en dessous
        Nu=0.36*(Re**0.55)*(Pr**(1/3))
        return(Nu)
        
    def Nusselt_plated(Re,Pr,alpha=0):
    
    #A l'extérieur des tubes.
        if (alpha<=5)  : Nu=0.989*((Pr)**(1/3))*((Re)**(0.330))
        if (5<alpha and alpha<=30 )  : Nu=0.140*((Pr)**(1/3))*((Re)**(0.729))
        if (30<alpha and alpha<=45 )  : Nu=0.338*((Pr)**(1/3))*((Re)**(0.659))
        if (45<alpha and alpha<=60 )  : Nu=0.455*((Pr)**(1/3))*((Re)**(0.660))
        if (60<alpha and alpha<=74 )  : Nu=0.525*((Pr)**(1/3))*((Re)**(0.640))
         
        return(Nu) 
        
    def f_factor(Re,alpha):
    
    #A l'extérieur des tubes.
        if (Re<3000 and alpha<5 )  : f=24/Re
        if (Re>3000 and alpha<5 )  : f=0.079*((Re)**(-0.25))
        if (40<Re and Re<500 and alpha>5 and alpha<=30 )  : f=23.33*((Re)**(-0.809))
        if (500<Re and Re<17000 and alpha>5 and alpha<=30 )  : f=0.557*((Re)**(-0.211))
        if (20<Re and Re<140 and alpha>=30 and alpha<60 )  : f=47.45*((Re)**(-0.680))
        if (140<Re and Re<4500 and alpha>=30 and alpha<60 )  : f=3.917*((Re)**(-0.175))
        if (40<Re and Re<180 and alpha>60 and alpha<90 )  : f=63.8*((Re)**(-0.809))
        if (180<Re and Re<700 and alpha>60 and alpha<90 )  : f=4.82*((Re)**(-0.312))

         
        return(f)
    
    def churchill(Re,alpha,gamma=1):
        beta=90-alpha
        p1=np.exp(-0.15705*beta)
        p2=np.pi*beta*gamma**2/3
        p3=np.exp(-np.pi*beta/(180*gamma**2))
        p4=(0.061+(0.69+np.tan(beta*np.pi/180))**(-2.63))*(1+(1-gamma)*0.9*beta**0.01)
        p5=1+beta/10
        A=(p4*np.log(p5/((7*p3/Re)**0.9+0.27e-5)))**16
        B=(37530*p1/Re)**16
        return 2*(((12+p2)/Re)**12+1/(A+B)**(3/2))**(1/12)
    
    def f(Re,alpha):
        
        if (alpha<=30):
            if(Re<10):
                return 50/Re
            elif(10<=Re and Re<=100):
                return 19.4/(Re**0.589)
            else:
                return 2.99/(Re**0.183)
        elif(30<alpha and alpha<=45):
            if(Re<15):
                return 47/Re
            elif(15<=Re and Re<=300):
                return 18.29/(Re**0.652)
            else:
                return 1.441/(Re**0.206)
        elif(45<alpha and alpha<=50):
            if(Re<20):
                return 34/Re
            elif(20<=Re and Re<=300):
                return 11.25/(Re**0.631)
            else:
                return 0.772/(Re**0.161)
        elif(50<alpha and alpha<=65):
            if(Re<40):
                return 24/Re
            elif(40<=alpha and alpha<=400):
                return 3.24/(Re**0.457)
            else:
                return 0.76/(Re**0.215)
        else:
            if(Re<50):
                return 24/Re
            elif(50<=Re and Re<=500):
                return 2.8/(Re**0.451)
            else:
                return 0.639/(Re**0.213)
    
    def h_calculation(Nu,k,D):
    
        h=(Nu*k)/D
        return(h)
        
    def DT_Co_courant(Tc_e,Tc_s,Tf_e,Tf_s):
    #Cas Co-Courant
        DT_e=Tc_e-Tf_e;
        DT_s=Tc_s-Tf_s;
        DT_m=(DT_s-DT_e)/math.log(DT_s/DT_e)
        #print("Différence log T en co-courant :",DT_m,"\n")
        return(DT_m)
        
    def DT_Contre_courant(Tc_e,Tc_s,Tf_e,Tf_s):
    #Cas Co-Courant
        DT_e=Tc_e-Tf_s
        DT_s=Tc_s-Tf_e
        if abs(DT_e-DT_s) <= 10**-7 : #suffisament proches pour les considérer comme égaux 
            DT_m = DT_e
        else :
            DT_m = (DT_s-DT_e)/np.log(DT_s/DT_e)
        # print("CONTRE-COURANT: "+str(Tc_e)+", "+str(Tf_s)+", "+str(Tc_s)+", "+str(Tf_e)+", ")
        #print("Différence log T en contre-courant :",DT_m,"\n")
        return(DT_m)
        
    def F_factor_1pC_2nT(Tc_e,Tc_s,Tf_e,Tf_s):
        R=(Tc_e-Tc_s)/(Tf_s-Tf_e)
        P=(Tf_s-Tf_e)/(Tc_e-Tf_e)
# 1 passe Calandre/ 2n passes de tubes
        F=(((np.sqrt(R**2 +1)/(R-1)))*np.log10((1-P)/(1-R*P)))/np.log10(((2/P)-1-R+np.sqrt(R**2 +1))/((2/P)-1-R-np.sqrt(R**2 +1)))
        if np.isfinite(F)==False:
            F=0.89
        if F>1:
            F=1
        return(F)
        
    def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,**to_excel_kwargs):
        """
        Append a DataFrame [df] to existing Excel file [filename]
        into [sheet_name] Sheet.
        If [filename] doesn't exist, then this function will create it.

        #Parameters:
        filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
                 df : dataframe to save to workbook
                 sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
                 startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
                 to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel`
                        [can be dictionary]

        Returns: None"""
        from openpyxl import load_workbook
        import pandas as pd
    # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

    # create a writer for this month and year
        writer = pd.ExcelWriter(filename, engine='openpyxl')

        try:
        # try to open an existing workbook
            writer.book = load_workbook(filename)
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
            if not startrow and sheet_name in writer.book.get_sheet_names():
                startrow = writer.book.get_sheet_by_name(sheet_name).max_row

        # copy existing sheets
            writer.sheets = dict(
                    (ws.title, ws) for ws in writer.book.worksheets)
        except FileNotFoundError:
        # file does not exist yet, we will create it
            pass

        if not startrow:
            startrow = 0

    # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
        writer.save()    
        
    def Blasius(Re):
    	F=0.316*Re**(-0.25)
    	return F
    
    def Colebrook(Re,k):
    	F=0.316*Re**(-0.25)
    	F2=1./(-2.*math.log10((2.51/(Re*math.sqrt(F)))+k*(1./3.71)))**2
    	i=0;
    	while abs(F2-F)>1e-4:
    		i+=1
    		F=F2
    		F2=1./(-2.*math.log10((2.51/(Re*math.sqrt(F)))+k*(1./3.71)))**2
    	return F2
    
    def Nikuradse(k):
    	F=1./(1.14-2.*math.log10(k))**2
    	return F